from fastapi import FastAPI, Form, UploadFile, File
from fastapi.responses import HTMLResponse
import uvicorn, os
from openpyxl import Workbook, load_workbook

app = FastAPI()

# ===================== DATA =====================

IT_JOBS = [
    "Python Developer","Java Developer","Full Stack Developer","Frontend Developer",
    "Backend Developer","AI / ML Engineer","Data Scientist","Data Analyst",
    "DevOps Engineer","Cloud Engineer","Cyber Security Analyst","QA / Tester",
    "Automation Engineer","Mobile App Developer","UI / UX Designer",
    "System Administrator","Network Engineer"
]

NON_IT_JOBS = [
    "HR Executive","HR Manager","Recruiter","Marketing Executive",
    "Digital Marketing Executive","Sales Executive","Business Development Executive",
    "Operations Executive","Customer Support","Office Admin","Accountant","Finance Executive"
]

QUALIFICATIONS = [
    "B.Tech","M.Tech","MCA","BCA","B.Sc","M.Sc","MBA",
    "Degree","Diploma","Intermediate","Others"
]

EXPERIENCE = list(range(0, 31))

QUESTIONS = {
    "Python Developer": [
        "Explain Python lists and tuples",
        "What is a dictionary in Python?",
        "What is OOP in Python?"
    ],
    "Java Developer": [
        "What is JVM?",
        "Difference between abstract class and interface",
        "Explain OOP concepts"
    ],
    "HR Executive": [
        "How do you handle employee conflict?",
        "What is recruitment process?",
        "How do you evaluate candidates?"
    ],
    "Marketing Executive": [
        "What is digital marketing?",
        "Explain SEO",
        "How do you generate leads?"
    ]
}

# ===================== EXCEL =====================

FILE_NAME = "applications.xlsx"

if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Name","Phone","Email","Experience","Qualification","Job Role",
        "Country","State","District","Area","AI Score","Result","Resume"
    ])
    wb.save(FILE_NAME)

def save_to_excel(data):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append(data)
    wb.save(FILE_NAME)

# ===================== AI EVAL =====================

def evaluate_answers(answers):
    score = 0
    for ans in answers:
        if len(ans.strip()) > 15:
            score += 1
    return score

# ===================== UI =====================

@app.get("/", response_class=HTMLResponse)
def home():
    return """
    <h2>AI Powered Job Application – Velvoro Software Solution</h2>
    <a href="/apply">Apply Now</a>
    """

@app.get("/apply", response_class=HTMLResponse)
def apply():
    job_options = "".join([f"<option>{j}</option>" for j in IT_JOBS + NON_IT_JOBS])
    qual_options = "".join([f"<option>{q}</option>" for q in QUALIFICATIONS])
    exp_options = "".join([f"<option>{e}</option>" for e in EXPERIENCE])

    return f"""
    <form method="post" action="/submit" enctype="multipart/form-data">
    <h3>AI Powered Job Application – Velvoro Software Solution</h3>

    <input name="name" placeholder="Full Name" required><br><br>
    <input name="phone" placeholder="Phone Number" required><br><br>
    <input name="email" placeholder="Email ID" required><br><br>

    <select name="experience">{exp_options}</select><br><br>
    <select name="qualification">{qual_options}</select><br><br>
    <select name="job_role">{job_options}</select><br><br>

    <input name="country" placeholder="Country"><br><br>
    <input name="state" placeholder="State"><br><br>
    <input name="district" placeholder="District"><br><br>
    <input name="area" placeholder="Area / Locality"><br><br>

    <input type="file" name="resume" required><br><br>

    <textarea name="q1" placeholder="Answer 1"></textarea><br><br>
    <textarea name="q2" placeholder="Answer 2"></textarea><br><br>
    <textarea name="q3" placeholder="Answer 3"></textarea><br><br>

    <button type="submit">Submit Application</button>
    </form>
    """

# ===================== SUBMIT =====================

@app.post("/submit", response_class=HTMLResponse)
async def submit(
    name: str = Form(...),
    phone: str = Form(...),
    email: str = Form(...),
    experience: int = Form(...),
    qualification: str = Form(...),
    job_role: str = Form(...),
    country: str = Form(...),
    state: str = Form(...),
    district: str = Form(...),
    area: str = Form(...),
    q1: str = Form(...),
    q2: str = Form(...),
    q3: str = Form(...),
    resume: UploadFile = File(...)
):
    answers = [q1, q2, q3]
    score = evaluate_answers(answers)
    result = "Qualified" if score >= 2 else "Not Qualified"

    resume_path = f"resumes_{resume.filename}"
    with open(resume_path, "wb") as f:
        f.write(await resume.read())

    save_to_excel([
        name, phone, email, experience, qualification, job_role,
        country, state, district, area, score, result, resume.filename
    ])

    return f"""
    <h2>Application Submitted</h2>
    <p>AI Result: <b>{result}</b></p>
    <p>Score: {score}/3</p>
    """

# ===================== RUN =====================

if __name__ == "__main__":
    uvicorn.run(
        "app:app",
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 8000))
    )
